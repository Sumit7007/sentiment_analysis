from transformers import AutoModelForSequenceClassification
from transformers import TFAutoModelForSequenceClassification
from transformers import AutoTokenizer, AutoConfig
import numpy as np
from scipy.special import softmax
import nltk
import re
import sys, os
import shutil
import math
import glob
import os
import openpyxl
import codecs
from tqdm import tqdm
import time

folder_path = r'C:\Users\sumit.shirole\Project\Sentiment_Project\input_cx'
excel_path = r'C:\Users\sumit.shirole\Project\Sentiment_Project\output_cx\ouput_cx.xlsx'
MODEL = f"cardiffnlp/twitter-xlm-roberta-base-sentiment"

class SentimentAnalysis():
    def __init__(self, folder_path, excel_path, MODEL):
        self.folder_path = folder_path
        self.excel_path = excel_path
        self.model = MODEL

    def extract_conversation(self,transcript):
        # remove speaker label and timestamp
        conversation = re.sub(r'SPEAKER_\d+\n\d+\.\d+--\d+\.\d+\n', '', transcript)
        # remove any remaining timestamps
        conversation = re.sub(r'\d+\.\d+--\d+\.\d+\n', '', conversation)
        # remove leading and trailing whitespace from each line
        conversation = '\n'.join([line.strip() for line in conversation.split('\n') if line.strip()])
        # remove newline characters
        conversation = conversation.replace('\n', ' ')
        return conversation


    def preprocess(self,text):
        text = text.lower()                             # Lowercasing          
        sentences = nltk.sent_tokenize(text)            # Sentence Tokenization
        return sentences

    def sentiment_model(self):
        try:
            shutil.rmtree("cardiffnlp")
        except:
            pass

        self.tokenizer = AutoTokenizer.from_pretrained(self.model)
        self.config = AutoConfig.from_pretrained(self.model)

        self.model = AutoModelForSequenceClassification.from_pretrained(self.model)
        # self.model.save_pretrained(self.model)
        # print('model loaded')

        return self.model
    
    def sentiment(self,token_sent):
        # print('inside sentiment')

        sent_list = []
        for i in range(0, len(token_sent), 6):
            text = token_sent[i:i + 6]
            text = ' '.join(text)
            encoded_input = self.tokenizer(text, return_tensors='pt', max_length=512, truncation='do_not_truncate')
            output = self.model(**encoded_input)
            scores = output[0][0].detach().numpy()
            scores = softmax(scores)

            j = 0
            ranking = np.argsort(scores)
            ranking = ranking[::-1]
            # list2.append(ranking)
            l = self.config.id2label[ranking[j]]
            # s = scores[ranking[j]]

            if l == 'positive':
                sent_list.append(5)
			    # sent_list.append(5)
            elif l == 'neutral':
                sent_list.append(3)
            elif l == 'negative':
                sent_list.append(1)

        return sum(sent_list)/len(sent_list), sent_list


    def normalize_value(self,value):
        """
        This function takes a value between (1,5) and returns a mormalized value between (-5,5).
        """
        normalized_value = ((value - 1) / 4) * 10 - 5
        return normalized_value
    

    def write_files_to_excel(self):
        # Create a new Excel workbook
        wb = openpyxl.Workbook()
        # Select the active sheet
        sheet = wb.active
        # Set the headers for the sheet
        sheet.cell(row=1, column=1, value='Filename')
        sheet.cell(row=1, column=2, value='Output')
        
        # Get a list of all the files in the folder
        files = os.listdir(self.folder_path)
        # Loop over files in the folder
        model = self.sentiment_model()
        
        for i, file in tqdm(enumerate(files),desc="Processing", ncols=80):
            time.sleep(0.1)
            # Read the contents of the file
            with codecs.open(os.path.join(self.folder_path, file), 'r',encoding='utf-8') as f:
                contents = f.read()
                text_after_label = self.extract_conversation(contents)
                tokenized_sentences = self.preprocess(text_after_label)
                var1 = self.sentiment(tokenized_sentences)
                var1_for_xsl = str(var1)

                # Normalize first value between the range of (-5,5)
                sentiment_score = self.normalize_value(var1[0])
                sentiment_score_for_xsl = str(sentiment_score)

            # Insert the filename and contents into a new row in the sheet
            sheet.cell(row=i+2, column=1, value=file)
            sheet.cell(row=i+2, column=2, value=var1_for_xsl)
            sheet.cell(row=i+2, column=3, value=sentiment_score_for_xsl)
        # Save the Excel file
        wb.save(self.excel_path)


if __name__ == '__main__':
    sent = SentimentAnalysis(folder_path,excel_path,MODEL)
    sent.write_files_to_excel()
